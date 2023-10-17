import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QLabel
from openpyxl import load_workbook
import os
import typing as t

import src.common.entities as entities
from src.backend import converter


class ExcelLoader:
    def __new__(cls):
        if not hasattr(cls, 'instance'):
            cls.instance = super(ExcelLoader, cls).__new__(cls)
        return cls.instance

    teams: t.List[entities.Team] = None

    def load_teams(self) -> t.List[entities.Team]:
        if self.teams:
            return self.teams
        cwd = os.getcwd()  # project_dir
        source_file = os.path.join(cwd, r'data\teams.xlsx')
        data = []
        try:
            wb = load_workbook(source_file)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
        except Exception as e:
            print(e)
            from src.gui.GUI import show_exception
            show_exception(str(e))

        self.teams = converter.convert_teams(data)
        return self.teams
